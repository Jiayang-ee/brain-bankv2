from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from faculty_spider_v3.models import SchoolSeed


SCHOOL_FIELDNAMES = [
    "rank",
    "school_name_en",
    "school_name_zh",
    "homepage_url",
    "difficulty_level",
    "crawl_status",
    "notes",
]


def read_school_seeds(path: str | Path, limit: int | None = None) -> list[SchoolSeed]:
    path = Path(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return read_school_seeds_excel(path, limit=limit)
    return read_school_seeds_csv(path, limit=limit)


def read_school_seeds_csv(path: str | Path, limit: int | None = None) -> list[SchoolSeed]:
    seeds: list[SchoolSeed] = []
    with Path(path).open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            name = _clean(row.get("school_name_en"))
            if not name:
                continue
            seeds.append(
                SchoolSeed(
                    rank=_int_or_none(row.get("rank")),
                    school_name_en=name,
                    school_name_zh=_clean(row.get("school_name_zh")),
                    homepage_url=_clean(row.get("homepage_url")),
                    difficulty_level=_int_or_none(row.get("difficulty_level")),
                    crawl_status=_clean(row.get("crawl_status")),
                    notes=_clean(row.get("notes")),
                )
            )
            if limit is not None and len(seeds) >= limit:
                break
    return seeds


def read_school_seeds_excel(path: str | Path, limit: int | None = None) -> list[SchoolSeed]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
    seeds: list[SchoolSeed] = []
    pending_rank: int | None = None
    pending_en = ""
    seen: set[str] = set()

    for row in worksheet.iter_rows(values_only=True):
        values = list(row[:6])
        rank = _int_or_none(values[0] if len(values) > 0 else None)
        name = _clean(values[1] if len(values) > 1 else None)
        homepage = _clean(values[2] if len(values) > 2 else None)
        difficulty = _int_or_none(values[3] if len(values) > 3 else None)
        status = _clean(values[4] if len(values) > 4 else None)
        notes = _clean(values[5] if len(values) > 5 else None)

        if not name or name in {"名称", "排名"} or name.startswith("爬取难度等级"):
            continue
        if rank is not None:
            pending_rank = rank
            pending_en = name
            continue
        if not pending_en:
            continue

        key = pending_en.casefold()
        if key in seen:
            continue
        seen.add(key)
        seeds.append(
            SchoolSeed(
                rank=pending_rank,
                school_name_en=pending_en,
                school_name_zh=name if name != pending_en else "",
                homepage_url=homepage,
                difficulty_level=difficulty,
                crawl_status=status,
                notes=notes,
            )
        )
        if limit is not None and len(seeds) >= limit:
            break
    return seeds


def write_school_seeds_csv(seeds: Iterable[SchoolSeed], path: str | Path) -> int:
    rows = list(seeds)
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SCHOOL_FIELDNAMES)
        writer.writeheader()
        for seed in rows:
            writer.writerow(seed.__dict__)
    return len(rows)


def _clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def _int_or_none(value: object) -> int | None:
    try:
        if value is None or str(value).strip() == "":
            return None
        return int(value)
    except (TypeError, ValueError):
        return None
